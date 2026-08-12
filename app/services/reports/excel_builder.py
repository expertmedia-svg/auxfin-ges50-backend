"""Construction du classeur Excel de rapprochement (section 16) : titres,
filtres, colonnes ajustees, ligne d'en-tete figee, couleurs sobres par statut."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

STATUS_COLORS = {
    "CONFORME": "DCFCE7",
    "ECART_DATE": "FEF3C7",
    "ECHEC_SYNCHRONISATION": "FEE2E2",
    "DECLARE_WHATSAPP_ABSENT_DASHBOARD": "FEE2E2",
    "DASHBOARD_SANS_PREUVE_WHATSAPP": "FEF3C7",
    "DOUBLON_WHATSAPP": "FDE68A",
    "DOUBLON_DASHBOARD": "FDE68A",
    "ID_ILLISIBLE": "E5E7EB",
    "DATE_ILLISIBLE": "E5E7EB",
    "SYNCHRONISATION_NON_CONFIRMEE": "E0E7FF",
    "VERIFICATION_MANUELLE": "E0E7FF",
    "ID_INCONNU": "FECACA",
    "MAUVAISE_APPLICATION": "FECACA",
}


def write_sheet(
    wb: Workbook, title: str, headers: list[str], rows: list[list], status_column: int | None = None
) -> Worksheet:
    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    if status_column is not None:
        for row_idx in range(2, len(rows) + 2):
            cell = ws.cell(row=row_idx, column=status_column)
            color = STATUS_COLORS.get(str(cell.value))
            if color:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows[:500]] or [0])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 50)

    return ws
